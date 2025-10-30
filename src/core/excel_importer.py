from openpyxl import load_workbook
from datetime import datetime
from typing import List, Tuple, Dict
from src.core.db_raw import Database
from src.utils.logger import logger


class ExcelImporter:
    
    def __init__(self):
        self.db = Database()
    
    def import_courses(self, file_path: str, department_id: int) -> Tuple[int, int, List[Dict]]:
        errors = []
        success_count = 0
        error_count = 0
        
        try:
            logger.info(f"📖 Ders listesi okunuyor: {file_path}")
            
            logger.info(f"  🗑️ Eski ders kayıtları temizleniyor (department_id: {department_id})...")
            
            old_courses = self.db.fetch_all(
                "SELECT id FROM courses WHERE department_id = ?",
                (department_id,)
            )
            
            if old_courses:
                old_course_ids = [c['id'] for c in old_courses]
                placeholders = ','.join('?' * len(old_course_ids))
                
                self.db.execute(
                    f"DELETE FROM student_courses WHERE course_id IN ({placeholders})",
                    tuple(old_course_ids)
                )
                
                self.db.execute(
                    "DELETE FROM courses WHERE department_id = ?",
                    (department_id,)
                )
                
                logger.info(f"  ✓ {len(old_courses)} eski ders kaydı silindi")
            else:
                logger.info(f"  ℹ️ Silinecek eski kayıt yok")
            
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                row_values = []
                for cell in row:
                    if cell:
                        val = str(cell).strip().upper()
                        val = val.replace('İ', 'I').replace('Ö', 'O').replace('Ü', 'U')
                        val = val.replace('Ğ', 'G').replace('Ş', 'S').replace('Ç', 'C')
                        row_values.append(val)
                    else:
                        row_values.append("")
                
                has_code = any("DERS" in val and "KOD" in val for val in row_values)
                has_name = any("DERSIN" in val and "ADI" in val for val in row_values)
                if has_code and has_name:
                    header_row = row_idx
                    logger.info(f"  ✓ Başlık satırı bulundu: Satır {header_row}")
                    break
            
            if not header_row:
                raise ValueError("Başlık satırı bulunamadı!")
            
            headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
            headers_normalized = []
            for h in headers:
                if h:
                    h_str = str(h).strip().upper()
                    h_str = h_str.replace('İ', 'I').replace('Ö', 'O').replace('Ü', 'U')
                    h_str = h_str.replace('Ğ', 'G').replace('Ş', 'S').replace('Ç', 'C')
                    headers_normalized.append(h_str)
                else:
                    headers_normalized.append("")
            
            code_col = None
            name_col = None
            instructor_col = None
            class_level_col = None
            
            for idx, header in enumerate(headers_normalized):
                if "DERS" in header and "KOD" in header:
                    code_col = idx
                elif "DERSIN" in header and "ADI" in header:
                    name_col = idx
                elif "DERSI" in header and "VEREN" in header:
                    instructor_col = idx
                elif "SINIF" in header or "SINIF" in header:
                    if class_level_col is None:
                        class_level_col = idx
            
            if code_col is None:
                raise ValueError("'DERS KODU' sütunu bulunamadı!")
            if name_col is None:
                raise ValueError("'DERSİN ADI' sütunu bulunamadı!")
            
            logger.info(f"  ✓ Sütunlar bulundu - Kod:{code_col}, Ad:{name_col}, Eğitmen:{instructor_col if instructor_col is not None else 'YOK'}")
            
            data_start_row = header_row + 1
            current_class_level = None
            is_elective = False
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                try:
                    if all(cell is None or str(cell).strip() == "" for cell in row):
                        continue
                    
                    course_code = str(row[code_col]).strip() if row[code_col] else ""
                    course_name = str(row[name_col]).strip() if row[name_col] else ""
                    instructor = str(row[instructor_col]).strip() if instructor_col is not None and row[instructor_col] else None
                    
                    if course_code.upper() in ["DERS KODU", "DERS KOD"] or course_name.upper() in ["DERSİN ADI", "DERSIN ADI", "DERS ADI"]:
                        logger.debug(f"    ⊗ Başlık satırı atlandı: Satır {row_idx}")
                        continue
                    
                    course_code_normalized = course_code.upper().replace('İ', 'I').replace('Ö', 'O').replace('Ü', 'U').replace('Ğ', 'G').replace('Ş', 'S').replace('Ç', 'C')
                    
                    if "SINIF" in course_code_normalized:
                        for i in range(1, 5):
                            if f"{i}." in course_code or str(i) in course_code:
                                current_class_level = i
                                is_elective = False
                                logger.debug(f"    📌 Sınıf seviyesi güncellendi: {i}. Sınıf (Satır {row_idx})")
                                break
                        continue
                    
                    if "SECMELI" in course_code_normalized and "DERS" in course_code_normalized:
                        is_elective = True
                        if current_class_level is None:
                            current_class_level = 3
                        logger.debug(f"    📌 Seçmeli ders bölümü başladı: {current_class_level}. Sınıf (Satır {row_idx})")
                        continue
                    
                    if "SECIMLIK" in course_code_normalized and "DERS" in course_code_normalized:
                        is_elective = True
                        current_class_level = 4
                        logger.debug(f"    📌 Seçimlik ders bölümü başladı: 4. Sınıf (Satır {row_idx})")
                        continue
                    
                    class_level_str = None
                    if class_level_col is not None and row[class_level_col]:
                        class_level_str = str(row[class_level_col]).strip()
                    
                    if not course_code:
                        errors.append({
                            "row": row_idx,
                            "error": "Ders kodu boş",
                            "data": {"name": course_name}
                        })
                        error_count += 1
                        continue
                    
                    if not course_name:
                        errors.append({
                            "row": row_idx,
                            "error": "Ders adı boş",
                            "data": {"code": course_code}
                        })
                        error_count += 1
                        continue
                    
                    final_class_level = current_class_level
                    if final_class_level is None:
                        final_class_level = self._extract_class_level(course_code)
                    
                    is_mandatory = not is_elective
                    
                    self.db.execute("""
                        INSERT INTO courses 
                        (code, name, instructor, department_id, class_level, is_mandatory, default_duration)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (course_code, course_name, instructor if instructor else None,
                          department_id, final_class_level, is_mandatory, 120))
                    logger.debug(f"    + Eklendi: {course_code} - {course_name} (Sınıf: {final_class_level}, {'Zorunlu' if is_mandatory else 'Seçmeli'})")
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append({
                        "row": row_idx,
                        "error": str(e),
                        "data": {"row": list(row)[:5]}
                    })
                    error_count += 1
                    logger.warning(f"    ✗ Satır {row_idx} hatası: {e}")
            
            logger.info(f"  ✅ Dersler içe aktarıldı: {success_count} başarılı, {error_count} hatalı")
            
            self._save_import_log(
                file_path, "courses", department_id,
                success_count, error_count, errors
            )
            
        except Exception as e:
            logger.error(f"  ❌ Kritik hata: {e}")
            raise
        
        return success_count, error_count, errors
    
    def import_students(self, file_path: str, department_id: int) -> Tuple[int, int, List[Dict]]:
        errors = []
        success_count = 0
        error_count = 0
        
        try:
            logger.info(f"👥 Öğrenci listesi okunuyor: {file_path}")
            
            logger.info(f"  🗑️ Eski öğrenci kayıtları temizleniyor (department_id: {department_id})...")
            
            old_students = self.db.fetch_all(
                "SELECT id FROM students WHERE department_id = ?",
                (department_id,)
            )
            
            if old_students:
                old_student_ids = [s['id'] for s in old_students]
                placeholders = ','.join('?' * len(old_student_ids))
                
                self.db.execute(
                    f"DELETE FROM student_courses WHERE student_id IN ({placeholders})",
                    tuple(old_student_ids)
                )
                
                self.db.execute(
                    "DELETE FROM students WHERE department_id = ?",
                    (department_id,)
                )
                
                logger.info(f"  ✓ {len(old_students)} eski öğrenci kaydı silindi")
            else:
                logger.info(f"  ℹ️ Silinecek eski kayıt yok")
            
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                row_values = [str(cell).strip() if cell else "" for cell in row]
                row_values_norm = [val.upper().replace("İ", "I").replace("Ö", "O").replace("Ğ", "G").replace("Ü", "U").replace("Ş", "S").replace("Ç", "C") for val in row_values]
                if any("OGRENCI" in val and "NO" in val for val in row_values_norm):
                    header_row = row_idx
                    logger.info(f"  ✓ Başlık satırı bulundu: Satır {header_row}")
                    break
            
            if not header_row:
                raise ValueError("Başlık satırı bulunamadı!")
            
            headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
            headers = [str(h).strip().upper().replace("İ", "I").replace("Ö", "O").replace("Ğ", "G").replace("Ü", "U").replace("Ş", "S").replace("Ç", "C") if h else "" for h in headers]
            
            student_no_col = None
            full_name_col = None
            class_level_col = None
            courses_col = None
            
            for idx, header in enumerate(headers):
                if "OGRENCI" in header and "NO" in header:
                    student_no_col = idx
                elif "AD" in header and "SOYAD" in header:
                    full_name_col = idx
                elif "SINIF" in header:
                    class_level_col = idx
                elif "DERS" in header and courses_col is None:
                    courses_col = idx
            
            if student_no_col is None:
                raise ValueError("'Öğrenci No' sütunu bulunamadı!")
            if full_name_col is None:
                raise ValueError("'Ad Soyad' sütunu bulunamadı!")
            
            logger.info(f"  ✓ Sütunlar bulundu - No:{student_no_col}, Ad:{full_name_col}, Sınıf:{class_level_col if class_level_col is not None else 'YOK'}, Ders:{courses_col if courses_col is not None else 'YOK'}")
            
            student_data = {}
            data_start_row = header_row + 1
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                try:
                    if all(cell is None or str(cell).strip() == "" for cell in row):
                        continue
                    
                    student_number = str(row[student_no_col]).strip() if row[student_no_col] else ""
                    full_name = str(row[full_name_col]).strip() if row[full_name_col] else ""
                    class_level = str(row[class_level_col]).strip() if class_level_col is not None and row[class_level_col] else None
                    course_code = str(row[courses_col]).strip() if courses_col is not None and row[courses_col] else None
                    
                    if not student_number or not full_name:
                        continue
                    
                    if student_number not in student_data:
                        student_data[student_number] = {
                            'name': full_name,
                            'class_level': class_level,
                            'courses': []
                        }
                    
                    if course_code and course_code not in student_data[student_number]['courses']:
                        student_data[student_number]['courses'].append(course_code)
                    
                except Exception as e:
                    logger.warning(f"    ✗ Satır {row_idx} okunamadı: {e}")
            
            logger.info(f"  ✓ {len(student_data)} benzersiz öğrenci bulundu")
            
            for student_number, data in student_data.items():
                try:
                    full_name = data['name']
                    class_level = data['class_level']
                    course_codes = data['courses']
                    
                    student_id = self.db.execute("""
                        INSERT INTO students (student_number, full_name, class_level, department_id)
                        VALUES (?, ?, ?, ?)
                    """, (student_number, full_name, class_level, department_id))
                    logger.debug(f"    + Eklendi: {student_number} - {full_name}")
                    
                    for course_code in course_codes:
                        course = self.db.fetch_one("""
                            SELECT * FROM courses 
                            WHERE code = ? AND department_id = ?
                        """, (course_code, department_id))
                        
                        if not course:
                            logger.warning(f"    ⚠ Ders bulunamadı: {course_code} (öğrenci: {student_number})")
                            continue
                        
                        course_id = dict(course)['id']
                        
                        self.db.execute("""
                            INSERT INTO student_courses (student_id, course_id)
                            VALUES (?, ?)
                        """, (student_id, course_id))
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append({
                        "row": student_number,
                        "error": str(e),
                        "data": {"number": student_number, "name": data.get('name', '')}
                    })
                    error_count += 1
                    logger.warning(f"    ✗ Öğrenci {student_number} hatası: {e}")
            
            logger.info(f"  ✅ Öğrenciler içe aktarıldı: {success_count} başarılı, {error_count} hatalı")
            
            self._save_import_log(
                file_path, "students", department_id,
                success_count, error_count, errors
            )
            
        except Exception as e:
            logger.error(f"  ❌ Kritik hata: {e}")
            raise
        
        return success_count, error_count, errors
    
    def _extract_class_level(self, course_code: str) -> int:
        import re
        match = re.search(r'\d+', course_code)
        if match:
            num = match.group()
            if len(num) >= 3:
                return int(num[0])
        return 1
    
    def _save_import_log(self, file_path: str, import_type: str,
                         department_id: int, success_count: int, error_count: int,
                         errors: List[Dict]):
        import json
        from pathlib import Path
        from datetime import datetime
        
        status = "success" if error_count == 0 else "partial" if success_count > 0 else "failed"
        total_rows = success_count + error_count
        
        self.db.execute("""
            INSERT INTO import_logs 
            (file_name, file_type, status, total_rows, success_rows, error_rows, error_details, imported_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            Path(file_path).name,
            import_type,
            status,
            total_rows,
            success_count,
            error_count,
            json.dumps(errors, ensure_ascii=False) if errors else None,
            datetime.now()
        ))
        
        logger.info(f"  📝 Import log kaydedildi: {Path(file_path).name} ({status})")
    
    def generate_error_csv(self, errors: List[Dict], output_path: str):
        import csv
        
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['Satır', 'Hata', 'Veri'])
            
            for error in errors:
                writer.writerow([
                    error.get('row', ''),
                    error.get('error', ''),
                    str(error.get('data', ''))
                ])
        
        logger.info(f"  📄 Hata raporu oluşturuldu: {output_path}")
