 
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QFileDialog, QComboBox
)
from PyQt6.QtCore import Qt
from datetime import datetime
import csv

from src.core.db_raw import Database
import logging

logger = logging.getLogger(__name__)


class ExamScheduleViewer(QWidget):
    
    def __init__(self, current_user, parent=None):
        super().__init__(parent)
        self.current_user = current_user
        self.db = Database()
        self.exam_schedules = []
        self.selected_schedule = None
        
        self.init_ui()
        self.load_schedules()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        title = QLabel("Sınav Programları")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
        layout.addWidget(title)
        
        select_layout = QHBoxLayout()
        select_layout.addWidget(QLabel("Program Seç:"))
        self.schedule_combo = QComboBox()
        self.schedule_combo.currentIndexChanged.connect(self.on_schedule_selected)
        select_layout.addWidget(self.schedule_combo)
        
        refresh_btn = QPushButton("Yenile")
        refresh_btn.clicked.connect(self.load_schedules)
        select_layout.addWidget(refresh_btn)
        
        export_btn = QPushButton("Excel'e Aktar")
        export_btn.clicked.connect(self.export_to_excel)
        select_layout.addWidget(export_btn)
        
        seating_btn = QPushButton("Oturma Planı")
        seating_btn.clicked.connect(self.show_seating_plan)
        seating_btn.setStyleSheet("background-color: #27ae60; color: white;")
        select_layout.addWidget(seating_btn)
        
        select_layout.addStretch()
        layout.addLayout(select_layout)
        
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Ders Kodu", "Ders Adı", "Tarih", "Saat", "Derslik(ler)", "Öğrenci Sayısı"
        ])

        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                gridline-color: #ddd;
                background-color: #fafbfc;
            }
            QTableWidget::item {
                padding: 8px;
                color: #2c3e50;
            }
            QTableWidget::item:alternate {
                background-color: #f8f9fa;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
        """)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        
        layout.addWidget(self.table)

        self.status_label = QLabel("Hazır")
        self.status_label.setStyleSheet("color: #7f8c8d; padding: 5px;")
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
    
    def load_schedules(self):
        
        try:
            self.schedule_combo.clear()
            rows = self.db.fetch_all("""
                SELECT * FROM exam_schedules 
                ORDER BY created_at DESC
            """)
            self.exam_schedules = [dict(row) for row in rows]
            
            for schedule in self.exam_schedules:
                status = "Tamamlandi" if schedule['is_finalized'] else "Taslak"
                self.schedule_combo.addItem(
                    f"{schedule['name']} ({status})",
                    schedule['id']
                )
            
            logger.info(f"[OK] {len(self.exam_schedules)} sinav programi yuklendi")
            
        except Exception as e:
            logger.error(f"Sinav programlari yukleme hatasi: {e}")
            QMessageBox.critical(self, "Hata", f"Programlar yüklenemedi:\n{str(e)}")
    
    def on_schedule_selected(self, index):
        
        if index < 0:
            return
        
        schedule_id = self.schedule_combo.itemData(index)
        if schedule_id:
            self.load_exam_schedule(schedule_id)
    
    def load_exam_schedule(self, schedule_id: int):
        
        try:
            schedule_row = self.db.fetch_one("SELECT * FROM exam_schedules WHERE id = ?", (schedule_id,))
            if not schedule_row:
                return
            self.selected_schedule = dict(schedule_row)
            
            exam_rows = self.db.fetch_all("""
                SELECT e.*, c.code as course_code, c.name as course_name
                FROM exams e
                JOIN courses c ON e.course_id = c.id
                WHERE e.schedule_id = ?
                ORDER BY e.exam_date, e.start_time
            """, (schedule_id,))
            exams = [dict(row) for row in exam_rows]
            
            self.table.setRowCount(len(exams))
            
            for row, exam in enumerate(exams):
                self.table.setItem(row, 0, QTableWidgetItem(exam['course_code']))
                self.table.setItem(row, 1, QTableWidgetItem(exam['course_name']))
                if exam.get('exam_date'):
                    if isinstance(exam['exam_date'], str):
                        date_str = datetime.fromisoformat(exam['exam_date']).strftime("%d.%m.%Y")
                    else:
                        date_str = exam['exam_date'].strftime("%d.%m.%Y")
                else:
                    date_str = "-"
                self.table.setItem(row, 2, QTableWidgetItem(date_str))
                if exam.get('start_time'):
                    if isinstance(exam['start_time'], str):
                        time_str = datetime.fromisoformat(f"2000-01-01 {exam['start_time']}").strftime("%H:%M")
                    else:
                        time_str = exam['start_time'].strftime("%H:%M")
                else:
                    time_str = "-"
                self.table.setItem(row, 3, QTableWidgetItem(time_str))
                session_rows = self.db.fetch_all("""
                    SELECT es.*, cl.code as classroom_code
                    FROM exam_sessions es
                    JOIN classrooms cl ON es.classroom_id = cl.id
                    WHERE es.exam_id = ?
                """, (exam['id'],))
                classrooms = [dict(sr)['classroom_code'] for sr in session_rows]
                classroom_str = ", ".join(classrooms) if classrooms else "-"
                self.table.setItem(row, 4, QTableWidgetItem(classroom_str))
                self.table.setItem(row, 5, QTableWidgetItem(str(exam.get('student_count', 0))))
            
            if self.selected_schedule.get('start_date'):
                if isinstance(self.selected_schedule['start_date'], str):
                    start_date_str = datetime.fromisoformat(self.selected_schedule['start_date']).strftime('%d.%m.%Y')
                else:
                    start_date_str = self.selected_schedule['start_date'].strftime('%d.%m.%Y')
            else:
                start_date_str = "-"
            
            if self.selected_schedule.get('end_date'):
                if isinstance(self.selected_schedule['end_date'], str):
                    end_date_str = datetime.fromisoformat(self.selected_schedule['end_date']).strftime('%d.%m.%Y')
                else:
                    end_date_str = self.selected_schedule['end_date'].strftime('%d.%m.%Y')
            else:
                end_date_str = "-"
            
            self.status_label.setText(
                f"[OK] {len(exams)} sinav yuklendi - {start_date_str} - {end_date_str}"
            )
            
            logger.info(f"[OK] {len(exams)} sinav goruntulendi")
            
        except Exception as e:
            logger.error(f"Sinav programi yukleme hatasi: {e}")
            QMessageBox.critical(self, "Hata", f"Program yüklenemedi:\n{str(e)}")
    
    def export_to_excel(self):
        
        if not self.selected_schedule:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce bir program seçin.")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            from collections import defaultdict

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                f"sinav_programi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return

            from src.core.db_raw import Database
            db = Database()
            
            exams_rows = db.fetch_all("""
                SELECT 
                    e.*,
                    c.code as course_code,
                    c.name as course_name,
                    c.instructor as course_instructor
                FROM exams e
                JOIN courses c ON e.course_id = c.id
                WHERE e.schedule_id = ?
                ORDER BY e.exam_date, e.start_time
            """, (self.selected_schedule['id'],))
            
            exams_by_date = defaultdict(list)
            for row in exams_rows:
                exam = dict(row)
                sessions_rows = db.fetch_all("""
                    SELECT 
                        cl.code as classroom_code,
                        es.allocated_seats
                    FROM exam_sessions es
                    JOIN classrooms cl ON es.classroom_id = cl.id
                    WHERE es.exam_id = ?
                """, (exam['id'],))
                classrooms_codes = [s['classroom_code'] for s in sessions_rows]
                exam['classrooms'] = "-".join(classrooms_codes) if classrooms_codes else "-"
                
                exams_by_date[exam['exam_date']].append(exam)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sınav Programı"
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            title_cell = ws.cell(row=1, column=1, value=self.selected_schedule['name'].upper())
            title_cell.font = Font(bold=True, size=13, color="000000")
            title_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.border = border
            ws.merge_cells('A1:E1')
            ws.row_dimensions[1].height = 25
            
            headers = ["Tarih", "Sınav Saati", "Ders Adı", "Öğretim Elemanı", "Derslik"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True, size=11, color="000000")
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            ws.row_dimensions[2].height = 30
            
            row_num = 3
            
            for date_str in sorted(exams_by_date.keys()):
                exams_on_date = exams_by_date[date_str]
                try:
                    date_obj = datetime.fromisoformat(date_str)
                    current_date_str = date_obj.strftime("%d.%m.%Y")
                except:
                    current_date_str = date_str
                
                exams_by_time = defaultdict(list)
                for exam in exams_on_date:
                    try:
                        time_obj = datetime.fromisoformat(exam['start_time']).time()
                        time_str = time_obj.strftime("%H.%M")
                    except:
                        time_str = str(exam['start_time'])[:5].replace(":", ".")
                    exams_by_time[time_str].append(exam)
                
                date_start_row = row_num
                
                for time_str in sorted(exams_by_time.keys()):
                    exams_at_time = exams_by_time[time_str]
                    
                    for exam in exams_at_time:
                        ws.cell(row=row_num, column=2, value=time_str)  # Saat
                        ws.cell(row=row_num, column=3, value=exam['course_name'])  # Ders Adı
                        ws.cell(row=row_num, column=4, value=exam.get('course_instructor', '-'))  # Öğretim Elemanı
                        ws.cell(row=row_num, column=5, value=exam['classrooms'])  # Derslik
                        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="center")
                        ws.cell(row=row_num, column=2).border = border
                        for col in [3, 4, 5]:
                            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="left", vertical="center")
                            ws.cell(row=row_num, column=col).border = border
                        
                        row_num += 1
                
                date_end_row = row_num - 1
                ws.cell(row=date_start_row, column=1, value=current_date_str)
                ws.cell(row=date_start_row, column=1).font = Font(bold=True, size=11)
                ws.cell(row=date_start_row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws.cell(row=date_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=date_start_row, column=1).border = border
                if date_end_row > date_start_row:
                    ws.merge_cells(f'A{date_start_row}:A{date_end_row}')
                    for r in range(date_start_row, date_end_row + 1):
                        ws.cell(row=r, column=1).border = border
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 45
            ws.column_dimensions['D'].width = 28
            ws.column_dimensions['E'].width = 20
            
            wb.save(file_path)
            
            total_exams = sum(len(exams) for exams in exams_by_date.values())
            QMessageBox.information(
                self,
                "Başarılı",
                f"✓ Sınav programı Excel'e aktarıldı!\n\n"
                f"📁 Dosya: {file_path}\n"
                f"📊 Toplam: {len(exams_by_date)} gün, {total_exams} sınav\n"
                f"📄 Format: Tek sayfa - Resmi üniversite şablonu"
            )
            logger.info(f"[OK] Sinav programi Excel'e export edildi: {file_path}")
            
        except ImportError:
            QMessageBox.critical(
                self, 
                "Hata", 
                "openpyxl kütüphanesi yüklü değil!\n\n"
                "Lütfen şu komutu çalıştırın:\npip install openpyxl"
            )
        except Exception as e:
            logger.error(f"Export hatasi: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Export başarısız:\n{str(e)}")
    
    def show_seating_plan(self):
        
        if not self.selected_schedule:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce bir program seçin!")
            return
        
        from src.ui.seating_plan_viewer import SeatingPlanViewer
        dialog = SeatingPlanViewer(self.selected_schedule, self.db, self)
        dialog.exec()
    
    def closeEvent(self, event):
        
        event.accept()
